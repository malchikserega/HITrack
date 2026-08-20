from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APIClient

from core.models import (
    Component,
    ComponentVersion,
    ComponentVersionVulnerability,
    Image,
    Vulnerability,
)


@override_settings(CACHES={
    'default': {'BACKEND': 'django.core.cache.backends.locmem.LocMemCache'},
})
class ReportGeneratorTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(User.objects.create_user('report-user'))
        self.image = Image.objects.create(name='registry.example.com/api:1.0.0')
        component = Component.objects.create(name='openssl', type='deb')
        component_version = ComponentVersion.objects.create(component=component, version='3.0.0')
        component_version.images.add(self.image)
        vulnerability = Vulnerability.objects.create(
            vulnerability_id='CVE-2024-0001', severity='HIGH',
        )
        ComponentVersionVulnerability.objects.create(
            component_version=component_version,
            vulnerability=vulnerability,
            fixable=True,
            fix='3.0.1',
        )

    def test_generates_xlsx_for_selected_image_with_uuid_primary_key(self):
        response = self.client.post(
            reverse('generate-report'),
            {'image_uuids': [str(self.image.uuid)]},
            format='json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], 'Image Name')
        self.assertEqual(rows[1], (
            self.image.name, 'openssl', 'deb', '3.0.0',
            'CVE-2024-0001', 'HIGH', '3.0.1',
        ))
